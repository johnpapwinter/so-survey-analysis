from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.borders import Border, Side, BORDER_THICK
from openpyxl.utils import get_column_letter


def export_to_excel(results: dict, country: str, year: int, category: str):

    # check if file exists, open it, else create it
    workbook = Workbook()
    # check if sheet exists for the country, if not create it
    workbook.create_sheet(country, 0)
    sheet = workbook.active
    sheet.title = country

    # create headers
    header_font = Font(bold=True, underline='single')
    sheet['A1'] = 'TYPE'
    sheet['A1'].font = header_font
    # this is returned by the helper function
    sheet['B1'] = '2023'
    sheet['A1'].font = header_font

    # go through the keys and assign value to a dictionary the key value and row number
    # if key does not exist, create it at the end of the column

    row_num = 2
    for key, value in results.items():
        sheet[f'A{row_num}'] = key
        sheet[f'B{row_num}'] = value
        row_num += 1

    # take the dictionary and assign the values to the rows

    workbook.save(f'{category}.xlsx')


def get_excel_column_for_year(year: int) -> str:
    if year == 2017:
        return 'B'
    elif year == 2018:
        return 'C'
    elif year == 2019:
        return 'D'
    elif year == 2020:
        return 'E'
    elif year == 2021:
        return 'F'
    elif year == 2022:
        return 'G'
    else:
        return 'H'


